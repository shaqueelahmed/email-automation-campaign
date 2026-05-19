"""
auto_reply.py — Module 3: Auto-Reply Automation
================================================
• Polls IMAP inbox (replies@yourdomain.com) every 60 seconds for UNSEEN messages
• Matches In-Reply-To / From header against ceo_data.xlsx
• Path A: Dispatches personalised auto-reply (Calendly link) for direct replies.
• Path B: Detects Calendly notifications and sends a final confirmation to the CEO.
• Logs every reply to Sheet 3 'Replies Log' in ceo_data.xlsx
"""

import os
import re
import email
import email.utils
import imaplib
import logging
import smtplib
import threading
import time
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

# ── IMAP (incoming) ───────────────────────────────────────────────────────
IMAP_HOST  = os.getenv("IMAP_HOST", "imap.gmail.com")
IMAP_PORT  = int(os.getenv("IMAP_PORT", "993"))
EMAIL_USER = os.getenv("IMAP_USER", "replies@yourdomain.com")
EMAIL_PASS = os.getenv("IMAP_PASSWORD", "")

# ── SMTP (outgoing) ───────────────────────────────────────────────────────
SMTP_HOST  = os.getenv("SMTP_HOST", "smtp.sendgrid.net")
SMTP_PORT  = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER  = os.getenv("SMTP_USER", "apikey")
SMTP_PASS  = os.getenv("SMTP_PASSWORD", "")

# ── Sender identity ───────────────────────────────────────────────────────
SENDER_NAME    = os.getenv("SENDER_NAME",    "Your Name")
SENDER_TITLE   = os.getenv("SENDER_TITLE",   "Head of Partnerships")
SENDER_COMPANY = os.getenv("SENDER_COMPANY", "Your Company")
SENDER_PHONE   = os.getenv("SENDER_PHONE",   "+1 000 000 0000")
SENDER_WEBSITE = os.getenv("SENDER_WEBSITE", "https://yourdomain.com")
MEETING_LINK   = os.getenv("CALENDAR_LINK",  "https://calendly.com/yourlink")
UNSUBSCRIBE_BASE = os.getenv("UNSUBSCRIBE_BASE", "https://yourdomain.com/unsubscribe")

# ── Files ─────────────────────────────────────────────────────────────────
EXCEL_FILE    = "ceo_data.xlsx"
TEMPLATE_FILE = "auto_reply_template.txt"
LOG_FILE      = "listener_log.txt"
LOG_SHEET     = "Replies Log"

POLL_INTERVAL = 60  # seconds between inbox checks

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)
_excel_lock = threading.Lock()

# ── Helpers ───────────────────────────────────────────────────────────────

def load_ceo_df() -> pd.DataFrame:
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name="CEO Master List")
    except Exception as exc:
        log.error("Cannot read CEO data: %s", exc)
        return pd.DataFrame()

def lookup_sender(from_email: str, df: pd.DataFrame) -> dict:
    """Return CEO row dict or sensible defaults."""
    if df.empty:
        return {}
    hit = df[df["Email Address"].str.lower() == from_email.lower()]
    return hit.iloc[0].to_dict() if not hit.empty else {}

def get_auto_reply(sender_email: str, sender_name: str) -> str:
    """Build personalised auto-reply body from template."""
    df  = load_ceo_df()
    ceo = lookup_sender(sender_email, df)

    first_name = str(ceo.get("Full Name", sender_name)).split()[0] if ceo else sender_name.split()[0]
    company    = str(ceo.get("Company Name", "your company")) if ceo else "your company"

    template = Path(TEMPLATE_FILE).read_text(encoding="utf-8")
    return (
        template
        .replace("{{first_name}}",    first_name)
        .replace("{{company}}",       company)
        .replace("{{meeting_link}}",  MEETING_LINK)
        .replace("{{sender_name}}",   SENDER_NAME)
        .replace("{{sender_title}}",  SENDER_TITLE)
        .replace("{{sender_company}}",SENDER_COMPANY)
        .replace("{{sender_phone}}",  SENDER_PHONE)
        .replace("{{sender_website}}",SENDER_WEBSITE)
        .replace("{{unsubscribe_link}}", f"{UNSUBSCRIBE_BASE}?email={sender_email}")
    )

def send_auto_reply(to_email: str, to_name: str, original_subject: str, body: str) -> bool:
    try:
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_USER
        msg["To"]      = to_email
        msg["Subject"] = f"Re: {original_subject} [Auto-Reply]" if not original_subject.startswith("Re:") else original_subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.ehlo()
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(EMAIL_USER, to_email, msg.as_string())
        log.info("[%s] Auto-replied to %s <%s>", datetime.now().strftime("%H:%M:%S"), to_name, to_email)
        return True
    except Exception as exc:
        log.error("Auto-reply FAILED for %s: %s", to_email, exc)
        return False

def append_replies_log(entry: dict):
    """Thread-safe append to 'Replies Log' sheet."""
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE)

        if LOG_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(LOG_SHEET)
            headers = [
                "Timestamp (UTC)", "From Email", "Sender Name",
                "CEO Name", "Company", "Original Subject",
                "Reply Sent At", "Status", "Reply Preview",
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center")
        else:
            ws = wb[LOG_SHEET]

        ws.append([
            entry["timestamp"],
            entry["from_email"],
            entry["sender_name"],
            entry["ceo_name"],
            entry["company"],
            entry["subject"],
            entry["reply_sent_at"],
            entry["status"],
            entry["reply_preview"],
        ])
        wb.save(EXCEL_FILE)

def process_message(raw_bytes: bytes, imap, uid: bytes):
    """Parse email and route to Direct Reply (Path A) or Calendly Booking (Path B)."""
    t_detect = time.time()

    msg         = email.message_from_bytes(raw_bytes)
    from_raw    = msg.get("From", "")
    sender_name, sender_email = email.utils.parseaddr(from_raw)
    subject     = msg.get("Subject", "(no subject)")
    
    # Extract email body for Regex parsing
    body_text = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                body_text = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                break
    else:
        body_text = msg.get_payload(decode=True).decode("utf-8", errors="ignore")

    log.info("New unseen email from %s <%s> | Subject: %s", sender_name, sender_email, subject)
    df = load_ceo_df()

    # ── PATH B: Booking Notification from Calendly ─────────────────────────
    if "calendly.com" in sender_email.lower():
        log.info("   ↳ Identified as a Calendly notification.")
        
        # Look for the exact string from your Calendly screenshot
        # ── THE UPGRADED REGEX FIX ───────────────────────────────────────────
        # This handles HTML tags, line breaks, and whitespace that might be 
        # hiding between "Invitee Email:" and the actual address.
        
        # Strip out all HTML tags just in case the body was parsed as HTML
        clean_body = re.sub(r"<[^>]+>", " ", body_text)
        
        # Look for the email with a much more forgiving regex
        extracted_email_match = re.search(r"Invitee Email:[\s\r\n]*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", clean_body, re.IGNORECASE)
        # ─────────────────────────────────────────────────────────────────────
        
        if not extracted_email_match:
            log.warning("   ↳ Could not find the Invitee Email in the Calendly body. Ignoring.")
            imap.store(uid, "+FLAGS", "\\Seen")
            return
            
        booked_ceo_email = extracted_email_match.group(1).strip()
        log.info("   ↳ Extracted booked email: %s", booked_ceo_email)
        
        booked_ceo = lookup_sender(booked_ceo_email, df)
        
        if not booked_ceo:
             log.info("   ↳ Ignored: The person who booked (%s) is not in our CEO Master List.", booked_ceo_email)
             imap.store(uid, "+FLAGS", "\\Seen")
             return
             
        # Send Confirmation Auto-Reply
        ceo_name = str(booked_ceo.get("Full Name", "there"))
        confirmation_body = (
            f"Hi {ceo_name.split()[0]},\n\n"
            f"I just saw your meeting confirmation come through. I'm really looking forward to our chat!\n\n"
            f"Best regards,\n"
            f"{SENDER_NAME}"
        )
        
        ok = send_auto_reply(booked_ceo_email, ceo_name, "Meeting Confirmed: Looking forward to our call", confirmation_body)
        
        append_replies_log({
            "timestamp":     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "from_email":    sender_email, # Calendly
            "sender_name":   "Calendly Notification",
            "ceo_name":      ceo_name,
            "company":       str(booked_ceo.get("Company Name", "Unknown")),
            "subject":       subject,
            "reply_sent_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "status":        "Sent (Booking Confirmation)" if ok else "Failed",
            "reply_preview": confirmation_body[:120].replace("\n", " ") + "…",
        })
        imap.store(uid, "+FLAGS", "\\Seen")
        return

    # ── PATH A: Direct Reply from a CEO ──────────────────────────────────
    ceo = lookup_sender(sender_email, df)
    
    if ceo:
        log.info("   ↳ Identified as a direct CEO reply. Sending Calendly link.")
        ceo_name = str(ceo.get("Full Name", sender_name))
        company  = str(ceo.get("Company Name", "Unknown"))
        
        reply_body = get_auto_reply(sender_email, sender_name)
        ok = send_auto_reply(sender_email, ceo_name, subject, reply_body)
        
        append_replies_log({
            "timestamp":     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "from_email":    sender_email,
            "sender_name":   sender_name,
            "ceo_name":      ceo_name,
            "company":       company,
            "subject":       subject,
            "reply_sent_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "status":        "Sent (Direct Reply)" if ok else "Failed",
            "reply_preview": reply_body[:120].replace("\n", " ") + "…",
        })
        
        elapsed = time.time() - t_detect
        log.info("   ↳ Reply dispatched in %.1f seconds.", elapsed)
        imap.store(uid, "+FLAGS", "\\Seen")
        return

    # ── PATH C: Ignore Everything Else (Spam, YouTube, etc.) ───────────────
    log.info("   ↳ Ignored: %s is not a CEO or Calendly notification.", sender_email)
    imap.store(uid, "+FLAGS", "\\Seen")

def listen_inbox():
    """Main polling loop — runs indefinitely."""
    seen_uids: set = set()
    log.info("Auto-reply listener started. Polling every %ds.", POLL_INTERVAL)

    while True:
        try:
            mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
            mail.login(EMAIL_USER, EMAIL_PASS)
            mail.select("inbox")

            _, data = mail.search(None, "UNSEEN")
            uids = data[0].split()

            if uids:
                log.info("Found %d unseen message(s).", len(uids))

            for uid in uids:
                if uid in seen_uids:
                    continue
                seen_uids.add(uid)
                _, raw = mail.fetch(uid, "(RFC822)")
                process_message(raw[0][1], mail, uid)

            mail.logout()

        except imaplib.IMAP4.error as exc:
            log.error("IMAP auth/connection error: %s", exc)
        except Exception as exc:
            log.error("Unexpected error: %s", exc)

        log.info("Sleeping %ds …", POLL_INTERVAL)
        time.sleep(POLL_INTERVAL)

if __name__ == "__main__":
    if not EMAIL_PASS:
        log.error("IMAP_PASSWORD not set in .env — aborting.")
    else:
        listen_inbox()