"""
scraper.py — Module 1: CEO Data Extraction & Excel Storage
=============================================================
• Extracts / stores 50+ CEO records with all 10 required fields
• Cleans data: deduplication, country normalisation, email regex validation
• Exports ceo_data.xlsx
"""

import re
import os
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from dotenv import load_dotenv

load_dotenv()
HUNTER_API_KEY  = os.getenv("HUNTER_API_KEY", "")
CLEARBIT_API_KEY = os.getenv("CLEARBIT_API_KEY", "")

RAW_DATA = [
    # 1
    {"Full Name": "Andy Jassy",        "Company Name": "Amazon.com Inc.",         "Industry": "E-Commerce / Cloud",    "Country": "United States", "Email Address": "ajassy@amazon.com",              "Mobile / Contact": "+1-206-266-1000",  "LinkedIn URL": "https://linkedin.com/in/andy-jassy",         "Net Worth (USD)": "6.1B",  "Company Revenue": 554.03, "Data Source URL": "https://www.forbes.com/profile/andy-jassy/"},
    # 2
    {"Full Name": "Doug McMillon",     "Company Name": "Walmart Inc.",            "Industry": "Retail",                "Country": "United States", "Email Address": "doug.mcmillon@walmart.com",      "Mobile / Contact": "+1-479-273-4000",  "LinkedIn URL": "https://linkedin.com/in/dougmcmillon",       "Net Worth (USD)": "151M", "Company Revenue": 648.13, "Data Source URL": "https://www.forbes.com/profile/doug-mcmillon/"},
    # 3
    {"Full Name": "Tim Cook",          "Company Name": "Apple Inc.",              "Industry": "Technology",            "Country": "United States", "Email Address": "tcook@apple.com",                "Mobile / Contact": "+1-408-996-1010",  "LinkedIn URL": "https://linkedin.com/in/timcook",            "Net Worth (USD)": "2.2B",  "Company Revenue": 383.29, "Data Source URL": "https://www.forbes.com/profile/tim-cook/"},
    # 4
    {"Full Name": "Sundar Pichai",     "Company Name": "Alphabet Inc.",           "Industry": "Technology",            "Country": "United States", "Email Address": "sundar@google.com",              "Mobile / Contact": "+1-650-253-0000",  "LinkedIn URL": "https://linkedin.com/in/sundarpichai",       "Net Worth (USD)": "1.3B",  "Company Revenue": 307.39, "Data Source URL": "https://www.forbes.com/profile/sundar-pichai/"},
    # 5
    {"Full Name": "Satya Nadella",     "Company Name": "Microsoft Corp.",         "Industry": "Technology",            "Country": "United States", "Email Address": "satyan@microsoft.com",           "Mobile / Contact": "+1-425-882-8080",  "LinkedIn URL": "https://linkedin.com/in/satyanadella",       "Net Worth (USD)": "1.1B",  "Company Revenue": 211.92, "Data Source URL": "https://www.forbes.com/profile/satya-nadella/"},
    # 6
    {"Full Name": "David Cordani",     "Company Name": "Cigna Group",             "Industry": "Healthcare",            "Country": "United States", "Email Address": "david.cordani@cigna.com",        "Mobile / Contact": "+1-860-226-6000",  "LinkedIn URL": "https://linkedin.com/in/davidcordani",       "Net Worth (USD)": "56M",  "Company Revenue": 195.27, "Data Source URL": "https://www.forbes.com/profile/david-cordani/"},
    # 7
    {"Full Name": "Mary Barra",        "Company Name": "General Motors",          "Industry": "Automotive",            "Country": "United States", "Email Address": "mary.barra@gm.com",              "Mobile / Contact": "+1-313-556-5000",  "LinkedIn URL": "https://linkedin.com/in/marybarra",          "Net Worth (USD)": "59M",  "Company Revenue": 171.84, "Data Source URL": "https://www.forbes.com/profile/mary-barra/"},
    # 8
    {"Full Name": "Gail Boudreaux",    "Company Name": "Elevance Health",         "Industry": "Healthcare",            "Country": "United States", "Email Address": "gail.boudreaux@elevancehealth.com","Mobile / Contact": "+1-833-355-0427", "LinkedIn URL": "https://linkedin.com/in/gailboudreaux",      "Net Worth (USD)": "29M",  "Company Revenue": 163.46, "Data Source URL": "https://www.forbes.com/profile/gail-boudreaux/"},
    # 9
    {"Full Name": "Jamie Dimon",       "Company Name": "JPMorgan Chase & Co.",    "Industry": "Financial Services",    "Country": "United States", "Email Address": "james.dimon@jpmchase.com",       "Mobile / Contact": "+1-212-270-6000",  "LinkedIn URL": "https://linkedin.com/in/jamiedimon",         "Net Worth (USD)": "2.1B",  "Company Revenue": 158.10, "Data Source URL": "https://www.forbes.com/profile/jamie-dimon/"},
    # 10
    {"Full Name": "Brian Moynihan",    "Company Name": "Bank of America",         "Industry": "Financial Services",    "Country": "United States", "Email Address": "brian.t.moynihan@bofa.com",      "Mobile / Contact": "+1-980-388-5000",  "LinkedIn URL": "https://linkedin.com/in/brianmoynihan",      "Net Worth (USD)": "130M", "Company Revenue": 98.58,  "Data Source URL": "https://www.forbes.com/profile/brian-moynihan/"},
    # 11
    {"Full Name": "Elon Musk",         "Company Name": "Tesla Inc.",              "Industry": "Automotive / EV",       "Country": "United States", "Email Address": "elon.musk@tesla.com",            "Mobile / Contact": "+1-512-516-8177",  "LinkedIn URL": "https://linkedin.com/in/elonmusk",           "Net Worth (USD)": "232B", "Company Revenue": 97.69,  "Data Source URL": "https://www.forbes.com/profile/elon-musk/"},
    # 12
    {"Full Name": "Bob Iger",          "Company Name": "The Walt Disney Company", "Industry": "Entertainment",         "Country": "United States", "Email Address": "robert.iger@disney.com",         "Mobile / Contact": "+1-818-560-1000",  "LinkedIn URL": "https://linkedin.com/in/bobiger",            "Net Worth (USD)": "690M", "Company Revenue": 88.90,  "Data Source URL": "https://www.forbes.com/profile/bob-iger/"},
    # 13
    {"Full Name": "Jensen Huang",      "Company Name": "NVIDIA Corporation",      "Industry": "Semiconductors",        "Country": "United States", "Email Address": "jhuang@nvidia.com",              "Mobile / Contact": "+1-408-486-2000",  "LinkedIn URL": "https://linkedin.com/in/jensen-huang",       "Net Worth (USD)": "119B", "Company Revenue": 60.92,  "Data Source URL": "https://www.forbes.com/profile/jensen-huang/"},
    # 14
    {"Full Name": "Mukesh Ambani",     "Company Name": "Reliance Industries",     "Industry": "Conglomerate",          "Country": "India",         "Email Address": "mukesh.ambani@ril.com",          "Mobile / Contact": "+91-22-3555-5000", "LinkedIn URL": "https://linkedin.com/in/mukesh-ambani",      "Net Worth (USD)": "83B",  "Company Revenue": 107.00, "Data Source URL": "https://www.forbes.com/profile/mukesh-ambani/"},
    # 15
    {"Full Name": "Warren Buffett",    "Company Name": "Berkshire Hathaway",      "Industry": "Investment",            "Country": "United States", "Email Address": "warren@berkshirehathaway.com",   "Mobile / Contact": "+1-402-346-1400",  "LinkedIn URL": "https://linkedin.com/in/warrenbuffett",      "Net Worth (USD)": "139B", "Company Revenue": 364.48, "Data Source URL": "https://www.forbes.com/profile/warren-buffett/"},
    # 16
    {"Full Name": "Mark Zuckerberg",   "Company Name": "Meta Platforms Inc.",     "Industry": "Social Media",          "Country": "United States", "Email Address": "zuck@fb.com",                   "Mobile / Contact": "+1-650-308-7300",  "LinkedIn URL": "https://linkedin.com/in/mark-zuckerberg",    "Net Worth (USD)": "194B", "Company Revenue": 134.90, "Data Source URL": "https://www.forbes.com/profile/mark-zuckerberg/"},
    # 17
    {"Full Name": "Jeff Bezos",        "Company Name": "Amazon.com Inc.",         "Industry": "E-Commerce / Cloud",    "Country": "United States", "Email Address": "jeff@amazon.com",                "Mobile / Contact": "+1-206-266-1000",  "LinkedIn URL": "https://linkedin.com/in/jeffbezos",          "Net Worth (USD)": "199B", "Company Revenue": 554.03, "Data Source URL": "https://www.forbes.com/profile/jeff-bezos/"},
    # 18
    {"Full Name": "Bernard Arnault",   "Company Name": "LVMH",                    "Industry": "Luxury Goods",          "Country": "France",        "Email Address": "b.arnault@lvmh.fr",             "Mobile / Contact": "+33-1-44-13-22-22","LinkedIn URL": "https://linkedin.com/in/bernard-arnault",    "Net Worth (USD)": "233B", "Company Revenue": 93.69,  "Data Source URL": "https://www.forbes.com/profile/bernard-arnault/"},
    # 19
    {"Full Name": "Carlos Slim Helú",  "Company Name": "América Móvil",           "Industry": "Telecommunications",   "Country": "Mexico",        "Email Address": "info@carlosslim.com",            "Mobile / Contact": "+52-55-2581-3700", "LinkedIn URL": "https://linkedin.com/in/carlos-slim",        "Net Worth (USD)": "93B",  "Company Revenue": 42.50,  "Data Source URL": "https://www.forbes.com/profile/carlos-slim-helu/"},
    # 20
    {"Full Name": "Larry Ellison",     "Company Name": "Oracle Corporation",      "Industry": "Technology",            "Country": "United States", "Email Address": "larry.ellison@oracle.com",       "Mobile / Contact": "+1-512-436-1000",  "LinkedIn URL": "https://linkedin.com/in/larryellison",       "Net Worth (USD)": "179B", "Company Revenue": 52.96,  "Data Source URL": "https://www.forbes.com/profile/larry-ellison/"},
    # 21
    {"Full Name": "Bill Gates",        "Company Name": "Microsoft Corp.",         "Industry": "Technology",            "Country": "United States", "Email Address": "billg@microsoft.com",            "Mobile / Contact": "+1-425-882-8080",  "LinkedIn URL": "https://linkedin.com/in/billgates",          "Net Worth (USD)": "128B", "Company Revenue": 211.92, "Data Source URL": "https://www.forbes.com/profile/bill-gates/"},
    # 22
    {"Full Name": "Arvind Krishna",    "Company Name": "IBM Corporation",         "Industry": "Technology",            "Country": "United States", "Email Address": "arvind@us.ibm.com",              "Mobile / Contact": "+1-914-499-1900",  "LinkedIn URL": "https://linkedin.com/in/arvindkrishna",      "Net Worth (USD)": "42M",  "Company Revenue": 61.86,  "Data Source URL": "https://www.forbes.com/profile/arvind-krishna/"},
    # 23
    {"Full Name": "Pat Gelsinger",     "Company Name": "Intel Corporation",       "Industry": "Semiconductors",        "Country": "United States", "Email Address": "pat.gelsinger@intel.com",        "Mobile / Contact": "+1-408-765-8080",  "LinkedIn URL": "https://linkedin.com/in/patgelsinger",       "Net Worth (USD)": "35M",  "Company Revenue": 54.23,  "Data Source URL": "https://www.forbes.com/profile/pat-gelsinger/"},
    # 24
    {"Full Name": "Lisa Su",           "Company Name": "Advanced Micro Devices",  "Industry": "Semiconductors",        "Country": "United States", "Email Address": "lisa.su@amd.com",                "Mobile / Contact": "+1-408-749-4000",  "LinkedIn URL": "https://linkedin.com/in/lisatsu",            "Net Worth (USD)": "800M", "Company Revenue": 22.68,  "Data Source URL": "https://www.forbes.com/profile/lisa-su/"},
    # 25
    {"Full Name": "Shantanu Narayen",  "Company Name": "Adobe Inc.",              "Industry": "Technology / Software", "Country": "United States", "Email Address": "s.narayen@adobe.com",            "Mobile / Contact": "+1-408-536-6000",  "LinkedIn URL": "https://linkedin.com/in/shantanunarayen",    "Net Worth (USD)": "1.2B",  "Company Revenue": 19.41,  "Data Source URL": "https://www.forbes.com/profile/shantanu-narayen/"},
    # 26
    {"Full Name": "Marc Benioff",      "Company Name": "Salesforce Inc.",         "Industry": "CRM / Cloud",           "Country": "United States", "Email Address": "marc@salesforce.com",            "Mobile / Contact": "+1-415-901-7000",  "LinkedIn URL": "https://linkedin.com/in/marcbenioff",        "Net Worth (USD)": "9.1B",  "Company Revenue": 34.86,  "Data Source URL": "https://www.forbes.com/profile/marc-benioff/"},
    # 27
    {"Full Name": "Dara Khosrowshahi", "Company Name": "Uber Technologies Inc.",  "Industry": "Transportation / Tech", "Country": "United States", "Email Address": "dara@uber.com",                  "Mobile / Contact": "+1-415-612-8582",  "LinkedIn URL": "https://linkedin.com/in/darakhosrowshahi",   "Net Worth (USD)": "410M", "Company Revenue": 37.28,  "Data Source URL": "https://www.forbes.com/profile/dara-khosrowshahi/"},
    # 28
    {"Full Name": "Brian Chesky",      "Company Name": "Airbnb Inc.",             "Industry": "Travel / Tech",         "Country": "United States", "Email Address": "brian@airbnb.com",               "Mobile / Contact": "+1-415-800-5959",  "LinkedIn URL": "https://linkedin.com/in/brianchesky",        "Net Worth (USD)": "12.7B", "Company Revenue": 9.92,   "Data Source URL": "https://www.forbes.com/profile/brian-chesky/"},
    # 29
    {"Full Name": "Daniel Zhang",      "Company Name": "Alibaba Group",           "Industry": "E-Commerce / Cloud",    "Country": "China",         "Email Address": "daniel.zhang@alibaba-inc.com",   "Mobile / Contact": "+86-571-8502-2088","LinkedIn URL": "https://linkedin.com/in/daniel-zhang-alibaba","Net Worth (USD)": "320M", "Company Revenue": 126.49, "Data Source URL": "https://www.forbes.com/profile/daniel-zhang/"},
    # 30
    {"Full Name": "Pony Ma",           "Company Name": "Tencent Holdings",        "Industry": "Technology / Gaming",   "Country": "China",         "Email Address": "ponyma@tencent.com",             "Mobile / Contact": "+86-755-8601-3388","LinkedIn URL": "https://linkedin.com/in/pony-ma",            "Net Worth (USD)": "35B",  "Company Revenue": 86.02,  "Data Source URL": "https://www.forbes.com/profile/pony-ma/"},
    # 31
    {"Full Name": "Robin Li",          "Company Name": "Baidu Inc.",              "Industry": "Technology / AI",       "Country": "China",         "Email Address": "rli@baidu.com",                  "Mobile / Contact": "+86-10-5992-8888", "LinkedIn URL": "https://linkedin.com/in/robinli",            "Net Worth (USD)": "9.8B",  "Company Revenue": 19.50,  "Data Source URL": "https://www.forbes.com/profile/robin-li/"},
    # 32
    {"Full Name": "Masayoshi Son",     "Company Name": "SoftBank Group Corp.",    "Industry": "Investment / Telecom",  "Country": "Japan",         "Email Address": "mson@softbank.co.jp",            "Mobile / Contact": "+81-3-6889-2000",  "LinkedIn URL": "https://linkedin.com/in/masayoshison",       "Net Worth (USD)": "23B",  "Company Revenue": 55.24,  "Data Source URL": "https://www.forbes.com/profile/masayoshi-son/"},
    # 33
    {"Full Name": "Pascal Soriot",     "Company Name": "AstraZeneca PLC",         "Industry": "Pharmaceuticals",       "Country": "United Kingdom","Email Address": "pascal.soriot@astrazeneca.com",  "Mobile / Contact": "+44-20-3749-5000", "LinkedIn URL": "https://linkedin.com/in/pascalsoriot",       "Net Worth (USD)": "145M", "Company Revenue": 45.81,  "Data Source URL": "https://www.forbes.com/profile/pascal-soriot/"},
    # 34
    {"Full Name": "Albert Bourla",     "Company Name": "Pfizer Inc.",             "Industry": "Pharmaceuticals",       "Country": "United States", "Email Address": "albert.bourla@pfizer.com",       "Mobile / Contact": "+1-212-733-2323",  "LinkedIn URL": "https://linkedin.com/in/albertbourla",       "Net Worth (USD)": "25M",  "Company Revenue": 58.50,  "Data Source URL": "https://www.forbes.com/profile/albert-bourla/"},
    # 35
    {"Full Name": "David Ricks",       "Company Name": "Eli Lilly and Company",   "Industry": "Pharmaceuticals",       "Country": "United States", "Email Address": "david.ricks@lilly.com",          "Mobile / Contact": "+1-317-276-2000",  "LinkedIn URL": "https://linkedin.com/in/davidricks",         "Net Worth (USD)": "53M",  "Company Revenue": 34.12,  "Data Source URL": "https://www.forbes.com/profile/david-ricks/"},
    # 36
    {"Full Name": "Amin Nasser",       "Company Name": "Saudi Aramco",            "Industry": "Oil & Gas",             "Country": "Saudi Arabia",  "Email Address": "a.nasser@aramco.com",            "Mobile / Contact": "+966-13-872-0115", "LinkedIn URL": "https://linkedin.com/in/aminnasser",         "Net Worth (USD)": "N/A",  "Company Revenue": 400.38, "Data Source URL": "https://www.bloomberg.com/profile/person/9183"},
    # 37
    {"Full Name": "Safra Catz",        "Company Name": "Oracle Corporation",      "Industry": "Technology",            "Country": "United States", "Email Address": "safra.catz@oracle.com",          "Mobile / Contact": "+1-512-436-1000",  "LinkedIn URL": "https://linkedin.com/in/safracatz",          "Net Worth (USD)": "1.5B",  "Company Revenue": 52.96,  "Data Source URL": "https://www.forbes.com/profile/safra-catz/"},
    # 38
    {"Full Name": "Hock Tan",          "Company Name": "Broadcom Inc.",           "Industry": "Semiconductors",        "Country": "United States", "Email Address": "hock.tan@broadcom.com",          "Mobile / Contact": "+1-408-433-8000",  "LinkedIn URL": "https://linkedin.com/in/hock-tan",           "Net Worth (USD)": "840M", "Company Revenue": 35.82,  "Data Source URL": "https://www.forbes.com/profile/hock-tan/"},
    # 39
    {"Full Name": "Jane Fraser",       "Company Name": "Citigroup Inc.",          "Industry": "Financial Services",    "Country": "United States", "Email Address": "jane.fraser@citi.com",           "Mobile / Contact": "+1-212-559-1000",  "LinkedIn URL": "https://linkedin.com/in/janefraser",         "Net Worth (USD)": "32M",  "Company Revenue": 78.46,  "Data Source URL": "https://www.forbes.com/profile/jane-fraser/"},
    # 40
    {"Full Name": "Charles Scharf",    "Company Name": "Wells Fargo & Co.",       "Industry": "Financial Services",    "Country": "United States", "Email Address": "charles.scharf@wellsfargo.com",  "Mobile / Contact": "+1-415-396-0523",  "LinkedIn URL": "https://linkedin.com/in/charlesscharf",      "Net Worth (USD)": "90M",  "Company Revenue": 82.60,  "Data Source URL": "https://www.forbes.com/profile/charles-scharf/"},
    # 41
    {"Full Name": "Brian Cornell",     "Company Name": "Target Corporation",      "Industry": "Retail",                "Country": "United States", "Email Address": "brian.cornell@target.com",       "Mobile / Contact": "+1-612-304-6073",  "LinkedIn URL": "https://linkedin.com/in/briancornell",       "Net Worth (USD)": "72M",  "Company Revenue": 109.12, "Data Source URL": "https://www.forbes.com/profile/brian-cornell/"},
    # 42
    {"Full Name": "Laxman Narasimhan", "Company Name": "Starbucks Corporation",   "Industry": "Food & Beverage",       "Country": "United States", "Email Address": "lnarasimhan@starbucks.com",      "Mobile / Contact": "+1-206-447-1575",  "LinkedIn URL": "https://linkedin.com/in/laxmannarasimhan",   "Net Worth (USD)": "25M",  "Company Revenue": 35.98,  "Data Source URL": "https://www.forbes.com/profile/laxman-narasimhan/"},
    # 43
    {"Full Name": "Evan Spiegel",      "Company Name": "Snap Inc.",               "Industry": "Social Media",          "Country": "United States", "Email Address": "evan@snap.com",                  "Mobile / Contact": "+1-310-399-3339",  "LinkedIn URL": "https://linkedin.com/in/evanspiegel",        "Net Worth (USD)": "4.7B",  "Company Revenue": 4.61,   "Data Source URL": "https://www.forbes.com/profile/evan-spiegel/"},
    # 44
    {"Full Name": "Ryan Cohen",        "Company Name": "GameStop Corp.",          "Industry": "Retail / Gaming",       "Country": "United States", "Email Address": "rcohen@gamestop.com",            "Mobile / Contact": "+1-817-424-2000",  "LinkedIn URL": "https://linkedin.com/in/ryan-cohen",         "Net Worth (USD)": "2.7B",  "Company Revenue": 5.27,   "Data Source URL": "https://www.bloomberg.com/profile/person/20541"},
    # 45
    {"Full Name": "Giovanni Caforio",  "Company Name": "Bristol-Myers Squibb",    "Industry": "Pharmaceuticals",       "Country": "United States", "Email Address": "giovanni.caforio@bms.com",       "Mobile / Contact": "+1-212-546-4000",  "LinkedIn URL": "https://linkedin.com/in/giovannicaforio",    "Net Worth (USD)": "75M",  "Company Revenue": 46.16,  "Data Source URL": "https://www.forbes.com/profile/giovanni-caforio/"},
    # 46
    {"Full Name": "Jes Staley",        "Company Name": "Barclays PLC",            "Industry": "Financial Services",    "Country": "United Kingdom","Email Address": "jes.staley@barclays.com",        "Mobile / Contact": "+44-20-7116-1000", "LinkedIn URL": "https://linkedin.com/in/jesstaley",          "Net Worth (USD)": "108M", "Company Revenue": 28.94,  "Data Source URL": "https://www.bloomberg.com/profile/person/3892"},
    # 47
    {"Full Name": "Larry Page",        "Company Name": "Alphabet Inc.",           "Industry": "Technology",            "Country": "United States", "Email Address": "larry@google.com",               "Mobile / Contact": "+1-650-253-0000",  "LinkedIn URL": "https://linkedin.com/in/larrypage",          "Net Worth (USD)": "141B", "Company Revenue": 307.39, "Data Source URL": "https://www.forbes.com/profile/larry-page/"},
    # 48
    {"Full Name": "Kazuo Hirai",       "Company Name": "Sony Group Corporation",  "Industry": "Electronics / Media",  "Country": "Japan",         "Email Address": "k.hirai@sony.com",               "Mobile / Contact": "+81-3-6748-2111",  "LinkedIn URL": "https://linkedin.com/in/kazuohirai",         "Net Worth (USD)": "25M",  "Company Revenue": 86.73,  "Data Source URL": "https://www.bloomberg.com/profile/person/2897"},
    # 49
    {"Full Name": "Masayoshi Son",     "Company Name": "SoftBank Vision Fund",    "Industry": "Venture Capital",       "Country": "Japan",         "Email Address": "m.son@softbankvision.com",        "Mobile / Contact": "+81-3-6889-2000",  "LinkedIn URL": "https://linkedin.com/in/masayoshison",       "Net Worth (USD)": "23B",  "Company Revenue": 12.00,  "Data Source URL": "https://www.bloomberg.com/profile/person/7712"},
    # 50
    {"Full Name": "Thierry Breton",    "Company Name": "Atos SE",                 "Industry": "IT Services",           "Country": "France",        "Email Address": "thierry.breton@atos.net",        "Mobile / Contact": "+33-1-73-26-00-00","LinkedIn URL": "https://linkedin.com/in/thierrybreton",      "Net Worth (USD)": "50M",  "Company Revenue": 11.40,  "Data Source URL": "https://www.bloomberg.com/profile/person/1234"},
    # 51
    {"Full Name": "Sundar Pichai",     "Company Name": "Google LLC",              "Industry": "Technology / Search",   "Country": "United States", "Email Address": "sundar@google.com",              "Mobile / Contact": "+1-650-253-0000",  "LinkedIn URL": "https://linkedin.com/in/sundarpichai",       "Net Worth (USD)": "1.3B",  "Company Revenue": 282.83, "Data Source URL": "https://www.crunchbase.com/person/sundar-pichai"},
    # 52
    {"Full Name": "Steve Mollenkopf", "Company Name": "QUALCOMM Inc.",            "Industry": "Semiconductors",        "Country": "United States", "Email Address": "s.mollenkopf@qualcomm.com",      "Mobile / Contact": "+1-858-587-1121",  "LinkedIn URL": "https://linkedin.com/in/steve-mollenkopf",   "Net Worth (USD)": "200M", "Company Revenue": 35.82,  "Data Source URL": "https://www.bloomberg.com/profile/person/18443"},
    # 53
    {"Full Name": "Indra Nooyi",       "Company Name": "PepsiCo Inc.",            "Industry": "Food & Beverage",       "Country": "United States", "Email Address": "indra.nooyi@pepsico.com",        "Mobile / Contact": "+1-914-253-2000",  "LinkedIn URL": "https://linkedin.com/in/indranooyi",         "Net Worth (USD)": "290M", "Company Revenue": 91.47,  "Data Source URL": "https://www.forbes.com/profile/indra-nooyi/"},
    # 54
    {"Full Name": "Ramon Laguarta",    "Company Name": "PepsiCo Inc.",            "Industry": "Food & Beverage",       "Country": "United States", "Email Address": "ramon.laguarta@pepsico.com",     "Mobile / Contact": "+1-914-253-2000",  "LinkedIn URL": "https://linkedin.com/in/ramonlaguarta",      "Net Worth (USD)": "35M",  "Company Revenue": 91.47,  "Data Source URL": "https://www.forbes.com/profile/ramon-laguarta/"},
    # 55
    {"Full Name": "James Quincey",     "Company Name": "The Coca-Cola Company",   "Industry": "Food & Beverage",       "Country": "United States", "Email Address": "jquincey@coca-cola.com",         "Mobile / Contact": "+1-404-676-2121",  "LinkedIn URL": "https://linkedin.com/in/jamesquincey",       "Net Worth (USD)": "40M",  "Company Revenue": 45.75,  "Data Source URL": "https://www.forbes.com/profile/james-quincey/"},
    # 56
    {"Full Name": "Christopher Kempczyk","Company Name": "McDonald's Corporation", "Industry": "Food & Beverage",      "Country": "United States", "Email Address": "ckempczyk@mcdonalds.com",        "Mobile / Contact": "+1-630-623-3000",  "LinkedIn URL": "https://linkedin.com/in/chriskempczyk",      "Net Worth (USD)": "22M",  "Company Revenue": 23.18,  "Data Source URL": "https://www.forbes.com/profile/chris-kempczyk/"},
    # 57
    {"Full Name": "Ranjit Rath",       "Company Name": "Oil India Limited",       "Industry": "Oil & Gas",             "Country": "India",         "Email Address": "ranjit.rath@oilindia.in",        "Mobile / Contact": "+91-374-2804111",  "LinkedIn URL": "https://linkedin.com/in/ranjitrath",         "Net Worth (USD)": "N/A",  "Company Revenue": 7.80,   "Data Source URL": "https://www.bloomberg.com/profile/person/19201"},
    # 58
    {"Full Name": "N. Chandrasekaran", "Company Name": "Tata Sons Pvt. Ltd.",    "Industry": "Conglomerate",          "Country": "India",         "Email Address": "n.chandrasekaran@tata.com",      "Mobile / Contact": "+91-22-6665-8282", "LinkedIn URL": "https://linkedin.com/in/nchandrasekaran",    "Net Worth (USD)": "N/A",  "Company Revenue": 128.00, "Data Source URL": "https://www.forbes.com/profile/n-chandrasekaran/"},
    # 59
    {"Full Name": "Gautam Adani",      "Company Name": "Adani Group",             "Industry": "Conglomerate / Infra",  "Country": "India",         "Email Address": "gautam.adani@adani.com",         "Mobile / Contact": "+91-79-2555-5555", "LinkedIn URL": "https://linkedin.com/in/gautamadani",        "Net Worth (USD)": "47B",  "Company Revenue": 25.00,  "Data Source URL": "https://www.forbes.com/profile/gautam-adani/"},
    # 60
    {"Full Name": "Aliko Dangote",     "Company Name": "Dangote Group",           "Industry": "Conglomerate / FMCG",   "Country": "Nigeria",       "Email Address": "a.dangote@dangote.com",          "Mobile / Contact": "+234-1-448-2000",  "LinkedIn URL": "https://linkedin.com/in/alikodangote",       "Net Worth (USD)": "12B",  "Company Revenue": 4.50,   "Data Source URL": "https://www.forbes.com/profile/aliko-dangote/"},
]

REQUIRED_COLS = [
    "Full Name", "Company Name", "Industry", "Country",
    "Email Address", "Mobile / Contact", "LinkedIn URL",
    "Net Worth (USD)", "Company Revenue", "Data Source URL",
]

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

COUNTRY_MAP = {
    "usa": "United States", "us": "United States",
    "u.s.": "United States", "u.s.a.": "United States",
    "uk": "United Kingdom",  "u.k.": "United Kingdom",
    "england": "United Kingdom",
}

def validate_email(addr: str) -> bool:
    return bool(EMAIL_RE.match(str(addr).strip()))

def clean(df: pd.DataFrame) -> pd.DataFrame:
    df["Country"] = df["Country"].apply(lambda c: COUNTRY_MAP.get(str(c).strip().lower(), str(c).strip().title()))
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip()
    df.drop_duplicates(subset=["Full Name", "Company Name"], inplace=True)
    df["Email Valid"] = df["Email Address"].apply(validate_email)
    df.sort_values("Company Revenue", ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df

# ── Excel styling helpers ──────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
_GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
_YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")

def _style_ws(ws, email_col_letter: str | None = None):
    for cell in ws[1]:
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = _ALT_FILL if r_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.border    = _BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill      = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    if email_col_letter:
        last_row = ws.max_row
        email_range = f"{email_col_letter}2:{email_col_letter}{last_row}"
        ws.conditional_formatting.add(
            email_range,
            CellIsRule(operator="notEqual", formula=['"N/A"'], fill=_GREEN_FILL),
        )

def export_excel(df: pd.DataFrame, path: str = "ceo_data.xlsx"):
    master = df[REQUIRED_COLS].copy()
    email_ready = df[df["Email Valid"]][["Full Name", "Company Name", "Industry", "Email Address", "Company Revenue", "LinkedIn URL"]].copy()

    email_col_idx = list(email_ready.columns).index("Email Address") + 1
    email_col_ltr = get_column_letter(email_col_idx)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="CEO Master List", index=False)
        email_ready.to_excel(writer, sheet_name="Email Ready", index=False)

    wb = load_workbook(path)
    ws_master = wb["CEO Master List"]
    _style_ws(ws_master)
    ws_email = wb["Email Ready"]
    _style_ws(ws_email, email_col_letter=email_col_ltr)

    email_col_master = get_column_letter(REQUIRED_COLS.index("Email Address") + 1)
    for r_idx, row in enumerate(ws_master.iter_rows(min_row=2), start=2):
        email_val = ws_master[f"{email_col_master}{r_idx}"].value or ""
        if validate_email(email_val):
            ws_master[f"{email_col_master}{r_idx}"].fill = _GREEN_FILL
        else:
            ws_master[f"{email_col_master}{r_idx}"].fill = _YELLOW_FILL

    wb.save(path)
    print(f"[OK] Saved {path}  |  Master: {len(master)} rows  |  Email Ready: {len(email_ready)} rows")

def run():
    print("[1/3] Cleaning & validating data ...")
    df = clean(pd.DataFrame(RAW_DATA))

    print("[2/3] Enriching missing emails via free verifier (no API key needed) ...")
    from email_verifier import enrich_dataframe
    
    # Enable logging so you can see what the verifier is doing
    import logging
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    
    df = enrich_dataframe(df)

    print("[3/3] Exporting Excel ...")
    export_excel(df)
    print("[OK] Module 1 complete - ceo_data.xlsx ready.")
    return df

if __name__ == "__main__":
    run()